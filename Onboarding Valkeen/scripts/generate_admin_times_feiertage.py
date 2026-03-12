#!/usr/bin/env python3
"""
Generiert Tempus Admin-Times Excel-Dateien mit Feiertagen.
- Nur Feiertags-Spalten (keine leeren Kalendertage)
- Aggregation Unit "Day"
- Werte: 8 Stunden Abwesenheit pro Feiertag (Time-basiert)
"""

from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Feiertage 2026 (Datum, Name) - national CH
CH_NATIONAL = [
    ("2026-01-01", "Neujahr"),
    ("2026-04-03", "Karfreitag"),
    ("2026-04-06", "Ostermontag"),
    ("2026-05-14", "Auffahrt"),
    ("2026-05-25", "Pfingstmontag"),
    ("2026-08-01", "Nationalfeiertag"),
    ("2026-12-25", "Weihnachten"),
    ("2026-12-26", "Stephanstag"),
]

# Zusätzliche Feiertage pro Schweizer Kanton (Kanton-Name -> Liste von (Datum, Name))
# Quelle: publicholidays.ch, feiertagskalender.ch
CH_KANTON_ZUSAETZLICH = {
    "Aargau": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-05-01", "Tag der Arbeit"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis")],
    "Appenzell Ausserrhoden": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-12-26", "Stephanstag")],
    "Appenzell Innerrhoden": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Basel-Landschaft": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-26", "Stephanstag")],
    "Basel-Stadt": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-26", "Stephanstag")],
    "Bern": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-26", "Stephanstag")],
    "Freiburg": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Genf": [("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-12-26", "Stephanstag")],
    "Glarus": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-04-09", "Näfelser Fahrt"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-26", "Stephanstag")],
    "Graubünden": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Jura": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-06-23", "Fest der Unabhängigkeit"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Luzern": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Neuenburg": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-12-26", "Stephanstag")],
    "Nidwalden": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Obwalden": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Schaffhausen": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-12-26", "Stephanstag")],
    "Schwyz": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Solothurn": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "St. Gallen": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Tessin": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-06-29", "St. Peter und Paul"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Thurgau": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-26", "Stephanstag")],
    "Uri": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Waadt": [("2026-01-02", "Berchtoldstag"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-12-26", "Stephanstag")],
    "Wallis": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    "Zug": [("2026-01-02", "Berchtoldstag"), ("2026-01-06", "Heilige Drei Könige"), ("2026-04-03", "Karfreitag"), ("2026-04-06", "Ostermontag"), ("2026-05-01", "Tag der Arbeit"), ("2026-05-25", "Pfingstmontag"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen"), ("2026-12-08", "Mariä Empfängnis"), ("2026-12-26", "Stephanstag")],
    # Zürich: nur gesetzliche Feiertage (zh.ch) – Berchtoldstag/Mariä Himmelfahrt nicht gesetzlich
    "Zürich": [("2026-05-01", "Tag der Arbeit")],
}

# Deutschland: Bundesweit + regional
DE_NATIONAL = [
    ("2026-01-01", "Neujahr"),
    ("2026-04-03", "Karfreitag"),
    ("2026-04-06", "Ostermontag"),
    ("2026-05-01", "Tag der Arbeit"),
    ("2026-05-14", "Christi Himmelfahrt"),
    ("2026-05-25", "Pfingstmontag"),
    ("2026-10-03", "Tag der Deutschen Einheit"),
    ("2026-12-25", "Weihnachten"),
    ("2026-12-26", "2. Weihnachtsfeiertag"),
]

# Österreich: gesetzliche Feiertage 2026 (bundesweit, feiertage-oesterreich.at)
# St. Josef (19.3.) nur in Kärnten/Steiermark/Tirol/Vorarlberg – hier weggelassen
AT_NATIONAL = [
    ("2026-01-01", "Neujahr"),
    ("2026-01-06", "Heilige Drei Könige"),
    ("2026-04-03", "Karfreitag"),
    ("2026-04-06", "Ostermontag"),
    ("2026-05-01", "Staatsfeiertag"),
    ("2026-05-14", "Christi Himmelfahrt"),
    ("2026-05-25", "Pfingstmontag"),
    ("2026-06-04", "Fronleichnam"),
    ("2026-08-15", "Mariä Himmelfahrt"),
    ("2026-10-26", "Nationalfeiertag"),
    ("2026-11-01", "Allerheiligen"),
    ("2026-12-08", "Mariä Empfängnis"),
    ("2026-12-25", "Christtag"),
    ("2026-12-26", "Stefanitag"),
]

DE_BUNDESLAND_ZUSAETZLICH = {
    "Baden-Württemberg": [("2026-01-06", "Heilige Drei Könige"), ("2026-06-04", "Fronleichnam"), ("2026-11-01", "Allerheiligen")],
    "Bayern": [("2026-01-06", "Heilige Drei Könige"), ("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen")],
    "Berlin": [("2026-03-08", "Internationaler Frauentag")],
    "Brandenburg": [("2026-10-31", "Reformationstag")],
    "Bremen": [("2026-10-31", "Reformationstag")],
    "Hamburg": [("2026-10-31", "Reformationstag")],
    "Hessen": [("2026-06-04", "Fronleichnam")],
    "Mecklenburg-Vorpommern": [("2026-03-08", "Internationaler Frauentag"), ("2026-10-31", "Reformationstag")],
    "Niedersachsen": [("2026-10-31", "Reformationstag")],
    "Nordrhein-Westfalen": [("2026-06-04", "Fronleichnam"), ("2026-11-01", "Allerheiligen")],
    "Rheinland-Pfalz": [("2026-06-04", "Fronleichnam"), ("2026-11-01", "Allerheiligen")],
    "Saarland": [("2026-06-04", "Fronleichnam"), ("2026-08-15", "Mariä Himmelfahrt"), ("2026-11-01", "Allerheiligen")],
    "Sachsen": [("2026-01-06", "Heilige Drei Könige"), ("2026-10-31", "Reformationstag")],
    "Sachsen-Anhalt": [("2026-01-06", "Heilige Drei Könige"), ("2026-10-31", "Reformationstag")],
    "Schleswig-Holstein": [("2026-10-31", "Reformationstag")],
    "Thüringen": [("2026-10-31", "Reformationstag")],
}


def get_holiday_dates(holidays: list[tuple[str, str]]) -> list[datetime]:
    """Nur Feiertags-Daten (Mo–Fr), sortiert. Keine leeren Kalendertage."""
    dates = []
    for date_str, _ in holidays:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        if dt.weekday() < 5:  # Nur Werktage
            dates.append(dt)
    return sorted(set(dates))


# Ressourcen für Valkeen (gemäss Anforderung)
RESOURCES = ["Manuel", "Marc", "Aayushi"]


def create_admin_times_excel(
    output_path: Path,
    region_name: str,
    holidays: list[tuple[str, str]],
    admin_time_type: str,
) -> None:
    """Erstellt eine Tempus Admin-Times Excel-Datei.
    Nur Feiertags-Spalten, 8 Stunden Abwesenheit pro Feiertag.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Admin Times"

    holiday_dates = get_holiday_dates(holidays)

    # Header: Resource Name, Aggregation Unit "Day", Admin Time Type, dann nur Feiertags-Spalten
    for col, val in enumerate(["Resource Name", "Aggregation Unit", "Admin Time Type"], 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = Font(bold=True)
    for col, d in enumerate(holiday_dates, 4):
        cell = ws.cell(row=1, column=col, value=d)
        cell.font = Font(bold=True)
        cell.number_format = "DD.MM.YYYY"

    # Datenzeilen: 8 Stunden Abwesenheit pro Feiertag
    hours_values = [8 for _ in holiday_dates]
    for row_idx, resource in enumerate(RESOURCES, 2):
        row_vals = [resource, "Day", admin_time_type] + hours_values
        for col, val in enumerate(row_vals, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # Anleitung als zweites Sheet
    ws_info = wb.create_sheet("Anleitung", 1)
    ws_info["A1"] = "Tempus Admin Times – Feiertage"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A2"] = f"Region: {region_name}"
    ws_info["A3"] = f"Admin Time Type: {admin_time_type}"
    ws_info["A4"] = f"Ressourcen: {', '.join(RESOURCES)}"
    ws_info["A5"] = ""
    ws_info["A6"] = "Import-Anleitung:"
    ws_info["A7"] = f"1. Admin-Time-Typ '{admin_time_type}' in Tempus anlegen (Admin Settings > Admin Times), falls noch nicht vorhanden."
    ws_info["A8"] = "2. Nur Feiertags-Spalten, 8 Stunden Abwesenheit pro Feiertag."
    ws_info["A9"] = "3. Ressourcennamen (Manuel, Marc, Aayushi) müssen exakt mit Tempus übereinstimmen."
    ws_info["A10"] = "4. Tempus: Admin Settings > Data Synchronization > Excel > Choose File > Synchronize."
    ws_info["A11"] = ""
    ws_info["A12"] = "Hinweis: Das Template-Sheet (Admin Times) muss das erste Tab sein."

    wb.save(output_path)


def main():
    script_dir = Path(__file__).parent
    output_dir = script_dir.parent / "docs" / "AdminTimes_Feiertage"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Schweiz: pro Kanton
    for kanton, zusaetzlich in CH_KANTON_ZUSAETZLICH.items():
        holidays = CH_NATIONAL + zusaetzlich
        # Duplikate entfernen (gleiches Datum)
        seen = set()
        unique = []
        for d, n in holidays:
            if d not in seen:
                seen.add(d)
                unique.append((d, n))
        admin_type = "Public Holidays Zurich" if kanton == "Zürich" else f"Public Holidays {kanton}"
        path = output_dir / f"Tempus_AdminTimes_Feiertage_CH_{kanton.replace(' ', '_')}.xlsx"
        create_admin_times_excel(path, f"Kanton {kanton}", unique, admin_type)
        print(f"Erstellt: {path.name}")

    # Deutschland: pro Bundesland
    for land, zusaetzlich in DE_BUNDESLAND_ZUSAETZLICH.items():
        holidays = DE_NATIONAL + zusaetzlich
        seen = set()
        unique = []
        for d, n in holidays:
            if d not in seen:
                seen.add(d)
                unique.append((d, n))
        admin_type = f"Public Holidays {land}"
        path = output_dir / f"Tempus_AdminTimes_Feiertage_DE_{land.replace(' ', '_')}.xlsx"
        create_admin_times_excel(path, land, unique, admin_type)
        print(f"Erstellt: {path.name}")

    # Österreich
    path = output_dir / "Tempus_AdminTimes_Feiertage_AT_Austria.xlsx"
    create_admin_times_excel(path, "Österreich", AT_NATIONAL, "Public Holidays Austria")
    print(f"Erstellt: {path.name}")

    print(f"\nAlle Dateien in: {output_dir}")


if __name__ == "__main__":
    main()

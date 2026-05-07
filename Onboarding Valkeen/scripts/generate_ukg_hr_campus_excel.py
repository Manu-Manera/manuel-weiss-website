#!/usr/bin/env python3
"""
UKG / HR Campus Demo – Assignments-Excel für Tempus Data Sync (Step 4)

Liest Aufgaben mit Startdatum 2026 aus Vorlage_Projektplan.xlsx (Sheet Projektplan)
und erzeugt eine Step-4-kompatible Datei für Projekt
  UKG Implementation — HR Campus Demo (API External Id: hr-campus-ukg-impl-demo-2026)

Verwendung:
  python generate_ukg_hr_campus_excel.py

Output:
  ../docs/UKG_HR_Campus_Assignments.xlsx

Import in Tempus: Admin Settings → Data Synchronization → Excel → Assignments
(Projektzeitraum sollte die importierten Monate abdecken.)

Benötigt: pip install openpyxl
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    raise SystemExit("Bitte installieren: pip install openpyxl")

REPO_ROOT = Path(__file__).resolve().parents[2]
VORLAGE = REPO_ROOT / "Vorlage_Projektplan.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "docs"
OUT_FILE = OUT_DIR / "UKG_HR_Campus_Assignments.xlsx"

PROJECT_NAME = "UKG Implementation — HR Campus Demo"
PROJECT_EXT_ID = "hr-campus-ukg-impl-demo-2026"

# Ressourcen (Namen wie in Tempus; External Id nur wo gesetzt)
ROWS_RESOURCE = [
    # Namen exakt wie in Tempus; External Id nur setzen, wenn in der Instanz gepflegt
    ("Manuel", "", 1.0),
    ("Giacomo Telesca", "", 0.55),
]

MONTH_HEADERS = [
    datetime(2026, m, 1, 0, 0) for m in range(1, 13)
]


def load_tasks_from_vorlage() -> list[tuple[str, datetime]]:
    """(task_name, start_date) für 2026, sinnvolle Labels."""
    wb = openpyxl.load_workbook(VORLAGE, data_only=True)
    ws = wb["Projektplan"]
    current_phase = ""
    out: list[tuple[str, datetime]] = []
    for r in range(14, 90):
        phase = ws.cell(r, 3).value
        product = ws.cell(r, 4).value
        task = ws.cell(r, 5).value
        start = ws.cell(r, 6).value
        if isinstance(phase, str) and phase.strip():
            current_phase = phase.strip()
        if not task or not isinstance(start, datetime):
            continue
        if start.year != 2026:
            continue
        t = task.strip() if isinstance(task, str) else str(task)
        prod = ""
        if isinstance(product, str) and product.strip():
            prod = f"{product.strip()} · "
        label = f"{current_phase + ' · ' if current_phase else ''}{prod}{t}"
        label = " ".join(label.split())
        out.append((label, start))
    return out


def build_assignment_row(
    project: str,
    proj_ext: str,
    res_name: str,
    res_ext: str,
    task: str,
    start: datetime,
    hours_scale: float,
) -> list:
    """Eine Plan-Zeile: Stunden nur im Startmonat (Rest 0)."""
    base_hours = round(32 * hours_scale)
    month_idx = start.month - 1
    allocs = [0] * 12
    allocs[month_idx] = max(8, base_hours)

    return [
        project,
        proj_ext,
        res_name,
        res_ext,
        task,
        "Hours",
        "Allocation",
        "Plan",
        "Month",
        "Merge",
        None,
        None,
        None,
        "Medium",
        "No",
    ] + allocs


def main() -> None:
    if not VORLAGE.exists():
        raise SystemExit(f"Vorlage fehlt: {VORLAGE}")

    tasks = load_tasks_from_vorlage()
    if not tasks:
        raise SystemExit("Keine 2026-Aufgaben in der Vorlage gefunden.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assignments"

    headers = [
        "Project",
        "Project API External Id",
        "Resource",
        "Resource API External Id",
        "Task",
        "Data Input",
        "Plan Type",
        "Allocation Type",
        "Time Period",
        "Import Behavior",
        "Id",
        "Action",
        "Project Allocation",
        "Priority",
        "Complete",
    ] + MONTH_HEADERS
    ws.append(headers)
    ws.row_dimensions[1].font = Font(bold=True)

    for task_name, start in tasks:
        for res_name, res_ext, factor in ROWS_RESOURCE:
            ws.append(
                build_assignment_row(
                    PROJECT_NAME,
                    PROJECT_EXT_ID,
                    res_name,
                    res_ext,
                    task_name,
                    start,
                    factor,
                )
            )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    print(f"✓ {OUT_FILE}")
    print(f"  Zeilen: {len(tasks)} Aufgaben × {len(ROWS_RESOURCE)} Ressourcen = {len(tasks) * len(ROWS_RESOURCE)} Assignment-Zeilen")
    print("  Import: Tempus → Admin Settings → Data Synchronization → Excel → Assignments")


if __name__ == "__main__":
    main()

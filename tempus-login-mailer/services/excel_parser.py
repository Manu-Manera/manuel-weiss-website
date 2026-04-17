import math
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


# E-Mail-Spalte: typische Überschriften (kein nacktes „adresse“ wegen Straßen-Spalten).
COLUMN_PATTERNS = {
    "email": re.compile(
        r"(e[-_\s]?mail|^\s*email\s*$|^\s*e_mail\s*$|^\s*mail\s*$|mail\s*adresse|"
        r"mail\s*address|e.?mail\s*adresse|empf.+mail|recipient|"
        r"work\s*e[-_]?mail|business\s*e[-_]?mail|primary\s*e[-_]?mail|contact\s*e[-_]?mail|"
        r"\bemail\s*address\b|\bemail\b)",
        re.IGNORECASE,
    ),
    "username": re.compile(r"(user\s?name|user\s?id|benutzer|benutzername|login.?id)", re.IGNORECASE),
    "password": re.compile(r"(pass\s?word|passwort|kennwort|pw|pwd)", re.IGNORECASE),
    "url": re.compile(r"(url|link|website|login.?url|zugang)", re.IGNORECASE),
    "name": re.compile(
        r"(^name$|full\s?name|display\s?name|anzeigename|client\s?name|kunde|kontakt|"
        r"empf[äa]nger|teilnehmer|"
        r"vor[\s\-/]*(?:und|&|\+|u\.)?[\s\-/]*nachname|"  # "Vor und Nachname", "Vor- und Nachname", "Vor-/Nachname"
        r"vor[\s\-/]*nachname|name\s*\(vorname)",
        re.IGNORECASE,
    ),
}

_VORNAME = re.compile(
    r"(^vorname$|^firstname$|^given\s?name$|^voornaam$|^pr[eé]nom$)",
    re.IGNORECASE,
)
_NACHNAME = re.compile(
    r"(^nachname$|^lastname$|^surname$|^family\s?name$|^achternaam$|^nom$)",
    re.IGNORECASE,
)

# E-Mail-Erkennung: zuerst ASCII-typisch, sonst weicher (Umlaute in lokalem Teil / IDN).
_EMAIL_LIKE_STRICT = re.compile(
    r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}",
)
# Kein \w für lokales Teil (Umlaute in Excel-Zellen sonst abgeschnitten).
_EMAIL_LIKE_LOOSE = re.compile(
    r"[^\s<>,;:\"']+@[^\s<>,;:\"']+\.[^\s<>,;:\"']{2,}",
)


def _cell_to_str(val) -> str:
    """Excel-Zellen zu sauberem Text (Zahlenformat, Leerzeichen)."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isfinite(val):
        if val == int(val):
            return str(int(val))
        s = str(val).strip()
        return s
    s = str(val).strip()
    return s


def _normalize_email(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    low = s.lower()
    if low.startswith("mailto:"):
        s = s[7:].strip()
    if "<" in s and ">" in s:
        m = re.search(r"<([^<>@\s]+@[^>]+)>", s)
        if m:
            s = m.group(1).strip()
    s = s.strip().strip('"').strip("'")
    if "@" not in s:
        return ""
    # Kein .search() mit „strict“: würde Teilstrings treffen (z. B. „ller@…“ in „müller@…“).
    if _EMAIL_LIKE_STRICT.fullmatch(s):
        return s
    if _EMAIL_LIKE_LOOSE.fullmatch(s):
        return s
    m = _EMAIL_LIKE_LOOSE.search(s)
    return m.group(0).strip() if m else ""


def _first_email_in_row(row: tuple) -> str:
    """Sucht in allen Zellen der Zeile nach einer E-Mail (Fallback, wenn die E-Mail-Spalte leer ist)."""
    if not row:
        return ""
    for cell in row:
        t = _normalize_email(_cell_to_str(cell))
        if t:
            return t
    return ""


def _detect_columns(headers: list[str], *, allow_empty_fallback: bool = True) -> dict[str, int | None]:
    """Map logical field names to column indices by matching header text."""
    keys = (
        "email",
        "username",
        "password",
        "url",
        "name",
        "vorname",
        "nachname",
    )
    mapping: dict[str, int | None] = {k: None for k in keys}

    for idx, header in enumerate(headers):
        header_str = _cell_to_str(header)
        if not header_str:
            continue
        if mapping["vorname"] is None and _VORNAME.search(header_str):
            mapping["vorname"] = idx
            continue
        if mapping["nachname"] is None and _NACHNAME.search(header_str):
            mapping["nachname"] = idx
            continue

    for idx, header in enumerate(headers):
        header_str = _cell_to_str(header)
        if not header_str:
            continue
        for field, pattern in COLUMN_PATTERNS.items():
            if mapping[field] is None and pattern.search(header_str):
                mapping[field] = idx
                break

    non_none = [v for v in mapping.values() if v is not None]
    if allow_empty_fallback and not non_none and len(headers) >= 1:
        mapping["email"] = 0

    return mapping


def _header_row_looks_like_data_row(headers: list[str], cmap: dict[str, int | None]) -> bool:
    """
    True, wenn die „Kopf“-Zeile eigentlich schon Daten ist (z. B. nur E-Mail-Adressen in jeder Zelle).
    """
    nonempty_i = [i for i, h in enumerate(headers) if _cell_to_str(h)]
    if not nonempty_i:
        return False
    n_email = sum(1 for i in nonempty_i if _normalize_email(_cell_to_str(headers[i])))
    other = sum(1 for k, v in cmap.items() if v is not None and k != "email")
    return other == 0 and n_email == len(nonempty_i) and n_email >= 1


def _header_row_score(headers: list[str]) -> int:
    """Höher = eher echte Kopfzeile (E-Mail-Spalte zählt am meisten)."""
    cmap = _detect_columns(headers, allow_empty_fallback=False)
    score = 0
    if cmap.get("email") is not None:
        score += 5
    for k in ("vorname", "nachname", "name", "username", "password", "url"):
        if cmap.get(k) is not None:
            score += 1
    return score


def _pick_header_row_index(rows: list[tuple], *, max_scan: int = 30) -> int:
    """
    Wählt die **erste** Zeile mit plausibler Kopfzeile (mehrzeilige Excel-Köpfe).

    Strenger als vorher: Keine spätere Datenzeile, die durch Zufall einen Score
    erreicht, überstimmt eine echte frühere Kopfzeile.
    """
    limit = min(max_scan, len(rows))
    for i in range(limit):
        hdr = [_cell_to_str(c) for c in rows[i]]
        if not any(h for h in hdr):
            continue
        s = _header_row_score(hdr)
        if s >= 2:
            return i
    for i in range(limit):
        hdr = [_cell_to_str(c) for c in rows[i]]
        if not any(h for h in hdr):
            continue
        if _header_row_score(hdr) >= 1 and not _first_email_in_row(rows[i]):
            return i
    return 0


_UMLAUT_MAP = str.maketrans({
    "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss",
    "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
})


def _translit(s: str) -> str:
    """Umlaute → ae/oe/ue/ss, Akzente entfernen, ASCII-ähnlich."""
    if not s:
        return ""
    s = s.translate(_UMLAUT_MAP)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s


def _split_firstlast(name: str) -> tuple[str, str]:
    """
    Zerlegt einen Namen in (vorname, nachname).
    Unterstützt „Nachname, Vorname“ (mit Komma) und „Vorname Nachname“.
    """
    n = (name or "").strip()
    if not n:
        return "", ""
    if "," in n:
        parts = [p.strip() for p in n.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            return parts[1], parts[0]
    tokens = n.split()
    if len(tokens) == 1:
        return tokens[0], ""
    return tokens[0], " ".join(tokens[1:])


def _normalize_name_cell(raw: str) -> str:
    """„Nachname, Vorname“ → „Vorname Nachname“; sonst unverändert."""
    vor, nach = _split_firstlast(raw)
    if vor and nach:
        return f"{vor} {nach}".strip()
    return (raw or "").strip()


# Kandidaten-Muster für die Bildung lokaler E-Mail-Teile (case-insensitive verglichen).
_LOCAL_PATTERNS: tuple[str, ...] = (
    "{vorname}.{nachname}",
    "{nachname}.{vorname}",
    "{v}.{nachname}",
    "{vorname}.{n}",
    "{vorname}{nachname}",
    "{nachname}{vorname}",
    "{vorname}_{nachname}",
    "{vorname}-{nachname}",
    "{v}{nachname}",
    "{vorname}",
    "{nachname}",
)


def _format_local(pattern: str, vorname: str, nachname: str) -> str:
    v = _translit(vorname).lower().replace(" ", "")
    n = _translit(nachname).lower().replace(" ", "")
    if not v and not n:
        return ""
    return pattern.format(
        vorname=v or "",
        nachname=n or "",
        v=(v[:1] if v else ""),
        n=(n[:1] if n else ""),
    )


def _learn_email_pattern(pairs: list[tuple[str, str]]) -> tuple[str, str] | None:
    """
    Versucht aus Paaren (name, email) das dominierende Muster + Domain zu lernen.
    Gibt (pattern, domain) zurück oder None.
    """
    votes: Counter[tuple[str, str]] = Counter()
    for name, email in pairs:
        if "@" not in email:
            continue
        local, _, domain = email.partition("@")
        local_l = local.lower()
        domain_l = domain.lower()
        vor, nach = _split_firstlast(name)
        if not (vor or nach):
            continue
        for pat in _LOCAL_PATTERNS:
            candidate = _format_local(pat, vor, nach)
            if candidate and candidate == local_l:
                votes[(pat, domain_l)] += 1
                break
    if not votes:
        return None
    (pat, dom), _ = votes.most_common(1)[0]
    return pat, dom


def _pick_dominant_domain(emails: list[str]) -> str:
    c: Counter[str] = Counter()
    for e in emails:
        _, _, dom = e.partition("@")
        dom = dom.strip().lower()
        if dom:
            c[dom] += 1
    return c.most_common(1)[0][0] if c else ""


@dataclass
class SheetReport:
    name: str
    total_rows: int = 0
    header_row_index: int = 0
    columns: dict[str, int | None] = field(default_factory=dict)
    data_rows: int = 0
    imported_with_email: int = 0
    imported_without_email: int = 0
    skipped_empty: int = 0
    inferred_emails: int = 0
    inferred_pattern: str = ""
    notes: list[str] = field(default_factory=list)


@dataclass
class ParseReport:
    file: str
    sheets: list[SheetReport] = field(default_factory=list)

    @property
    def total_imported(self) -> int:
        return sum(s.imported_with_email + s.imported_without_email for s in self.sheets)

    @property
    def total_without_email(self) -> int:
        return sum(s.imported_without_email for s in self.sheets)

    @property
    def total_skipped_empty(self) -> int:
        return sum(s.skipped_empty for s in self.sheets)

    def human_summary(self) -> str:
        lines = [f"Datei: {self.file}"]
        for s in self.sheets:
            cols = ", ".join(f"{k}={v}" for k, v in s.columns.items() if v is not None) or "— keine erkannt —"
            lines.append(
                f"Blatt „{s.name}“: {s.total_rows} Zeilen, Kopf in Zeile {s.header_row_index + 1}, "
                f"Spalten: {cols}"
            )
            lines.append(
                f"  → importiert: {s.imported_with_email + s.imported_without_email} "
                f"({s.imported_without_email} ohne E-Mail), "
                f"leere Zeilen übersprungen: {s.skipped_empty}"
            )
            if s.inferred_emails:
                lines.append(
                    f"  ↳ {s.inferred_emails} E-Mail(s) aus Muster ergänzt: {s.inferred_pattern}"
                )
            for n in s.notes:
                lines.append(f"  Hinweis: {n}")
        return "\n".join(lines)


def _parse_sheet_rows(
    rows: list[tuple],
    *,
    sheet_name: str = "",
) -> tuple[list[dict[str, str]], SheetReport]:
    report = SheetReport(name=sheet_name, total_rows=len(rows))

    if not rows:
        return [], report

    hi = _pick_header_row_index(rows)
    report.header_row_index = hi
    headers = [_cell_to_str(c) for c in rows[hi]]
    col_map = _detect_columns(headers)

    # Sonderfall: erste Zeile enthält schon @-Adressen (kein echter Kopf) → alles als Daten.
    header_row_is_data = False
    if col_map["email"] is None:
        for idx, cell in enumerate(rows[hi]):
            if cell and "@" in _cell_to_str(cell):
                col_map["email"] = idx
                header_row_is_data = True
                break
    if not header_row_is_data and _header_row_looks_like_data_row(headers, col_map):
        header_row_is_data = True

    if header_row_is_data:
        data_start = hi
        report.notes.append("Erste Zeile enthält bereits Daten — kein separater Kopf.")
    else:
        data_start = hi + 1

    report.columns = dict(col_map)
    data_rows = rows[data_start:]
    report.data_rows = len(data_rows)

    def _col(row: tuple, idx: int | None) -> str:
        if idx is None or idx < 0:
            return ""
        if idx >= len(row):
            return ""
        return _cell_to_str(row[idx])

    results: list[dict[str, str]] = []
    for row in data_rows:
        if not row or all(c is None or _cell_to_str(c) == "" for c in row):
            report.skipped_empty += 1
            continue

        entry: dict[str, str] = {}
        for fname in ("email", "username", "password", "url", "name"):
            entry[fname] = _col(row, col_map.get(fname))

        vor = _col(row, col_map.get("vorname"))
        nach = _col(row, col_map.get("nachname"))
        combined = f"{vor} {nach}".strip()
        single_name = _normalize_name_cell(entry.get("name") or "")
        entry["name"] = combined or single_name

        raw_email_cell = entry.get("email", "")
        email = _normalize_email(raw_email_cell)
        if not email:
            email = _first_email_in_row(row)

        # Notfall: Wenn in der „E-Mail“-Spalte nur ein Name wie „Nachname, Vorname“ steht
        # (keine echte Adresse), nutzen wir ihn als Namen, falls sonst keiner da ist.
        if not email and raw_email_cell and not entry["name"]:
            maybe_name = _normalize_name_cell(raw_email_cell)
            if maybe_name and " " in maybe_name and "@" not in maybe_name:
                entry["name"] = maybe_name

        entry["email"] = email
        results.append(entry)

    # Muster aus vorhandenen Paaren (Name, E-Mail) lernen und fehlende Adressen ableiten.
    pairs = [(r["name"], r["email"]) for r in results if r["name"] and r["email"]]
    learned = _learn_email_pattern(pairs)
    if not learned:
        domain = _pick_dominant_domain([r["email"] for r in results if r["email"]])
        if domain:
            learned = ("{vorname}.{nachname}", domain)

    if learned:
        pattern, domain = learned
        report.inferred_pattern = f"{pattern}@{domain}"
        for r in results:
            if r["email"]:
                continue
            vorname, nachname = _split_firstlast(r.get("name", ""))
            if not (vorname and nachname):
                continue
            local = _format_local(pattern, vorname, nachname)
            if not local:
                continue
            r["email"] = f"{local}@{domain}"
            report.inferred_emails += 1

    report.imported_with_email = sum(1 for r in results if r["email"])
    report.imported_without_email = sum(1 for r in results if not r["email"])

    return results, report


def parse_excel(file_path: str | Path) -> list[dict[str, str]]:
    """
    Parse an Excel file and return a list of dicts with keys:
    email, name, url, username, password.

    - Erkennt die Kopfzeile auch bei mehrzeiligem Kopf oben im Blatt.
    - Liest **alle Arbeitsblätter** nacheinander (Reihenfolge wie in Excel).
    - **Keine** stillschweigende Deduplizierung nach E-Mail.
    - **Zeilen ohne erkennbare E-Mail werden ebenfalls übernommen** (mit leerem
      E-Mail-Feld), damit in der App nichts „still verschwindet“. Beim Versand
      werden nur Zeilen mit E-Mail berücksichtigt.
    """
    rows, _ = parse_excel_with_report(file_path)
    return rows


def _mailto_from_hyperlink(cell) -> str:
    """Extrahiert eine E-Mail-Adresse aus einem Zellen-Hyperlink (z. B. mailto:…)."""
    hl = getattr(cell, "hyperlink", None)
    if hl is None:
        return ""
    target = getattr(hl, "target", "") or getattr(hl, "location", "") or ""
    target = str(target).strip()
    if not target:
        return ""
    low = target.lower()
    if low.startswith("mailto:"):
        target = target[7:]
        target = target.split("?", 1)[0]
    return _normalize_email(target)


def _resolve_rows_with_hyperlinks(ws) -> list[tuple]:
    """
    Liest ein Arbeitsblatt als Liste von Zeilen-Tupeln (Strings). Enthält eine Zelle
    einen `mailto:`-Hyperlink, wird die **Hyperlink-Adresse** statt des Anzeigetextes
    verwendet. Wird ein Hyperlink auf eine andere URL gesetzt, aber die Zelle enthält
    Text wie „Nachname, Vorname“, wird die Adresse zusätzlich in eine virtuelle zweite
    Zelle gehängt, damit der E-Mail-Fallback sie pro Zeile finden kann.
    """
    out: list[tuple] = []
    for row in ws.iter_rows():
        cells: list = []
        extra_email = ""
        for cell in row:
            value = cell.value
            mail = _mailto_from_hyperlink(cell)
            if mail:
                existing = _normalize_email(_cell_to_str(value))
                if existing:
                    cells.append(value)
                else:
                    # Anzeigetext behalten, damit der Name erhalten bleibt;
                    # die E-Mail landet per extra_email in der Zeile.
                    cells.append(value if value not in (None, "") else mail)
                    if not existing:
                        extra_email = mail
            else:
                cells.append(value)
        if extra_email:
            cells.append(extra_email)
        out.append(tuple(cells))
    return out


def parse_excel_with_report(file_path: str | Path) -> tuple[list[dict[str, str]], ParseReport]:
    # Kein read_only: sonst fehlen die Hyperlinks („mailto:…“), die in
    # Excel-Exporten aus Outlook/Adressbuch die eigentliche Adresse tragen.
    wb = load_workbook(str(file_path), data_only=True)
    report = ParseReport(file=str(file_path))
    merged: list[dict[str, str]] = []

    try:
        for ws in wb.worksheets:
            rows = _resolve_rows_with_hyperlinks(ws)
            parsed, sheet_report = _parse_sheet_rows(rows, sheet_name=ws.title)
            report.sheets.append(sheet_report)
            merged.extend(parsed)
    finally:
        wb.close()

    return merged, report
